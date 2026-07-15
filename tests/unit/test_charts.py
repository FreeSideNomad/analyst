"""Feature 014 — saved charts (domain + repository lifecycle) and exports."""

from __future__ import annotations

import pytest

from analyst.api.repository import StoreRepository
from analyst.domain.charts import (
    ChartDataGoneError,
    SavedChart,
    UnknownChartError,
    chart_id_for,
)

SALES = "region,amount\nEast,10\nEast,20\nEast,30\nWest,40\nWest,50\n"
TOTALS_SQL = 'SELECT region, SUM(amount) AS total FROM "sales.csv" GROUP BY region ORDER BY total DESC'


def _repo(tmp_path) -> StoreRepository:
    return StoreRepository(str(tmp_path / "data"))


def _repo_with_sales(tmp_path) -> tuple[StoreRepository, str]:
    repo = _repo(tmp_path)
    (rec,) = repo.ingest("sales.csv", SALES.encode())
    return repo, rec.name


def _save(repo, name="Revenue by region") -> SavedChart:
    return repo.save_chart(
        name=name,
        question="What is the total amount by region?",
        sql=TOTALS_SQL,
        chart_type="bar",
        title="Total amount by region",
        datasets=["sales.csv"],
    )


# --------------------------------------------------------------------------- #
# Domain
# --------------------------------------------------------------------------- #
def test_chart_ids_are_stable_slugs_and_collision_free():
    assert chart_id_for("Revenue by region", set()) == "revenue-by-region"
    assert (
        chart_id_for("Revenue by region", {"revenue-by-region"})
        == "revenue-by-region-2"
    )


# --------------------------------------------------------------------------- #
# Repository CRUD + persistence
# --------------------------------------------------------------------------- #
def test_save_and_list_charts(tmp_path):
    repo, _ = _repo_with_sales(tmp_path)
    chart = _save(repo)
    assert chart.chart_id == "revenue-by-region"
    assert [c.name for c in repo.charts()] == ["Revenue by region"]


def test_charts_survive_a_restart(tmp_path):
    repo, _ = _repo_with_sales(tmp_path)
    _save(repo)
    reborn = _repo(tmp_path)
    assert [c.name for c in reborn.charts()] == ["Revenue by region"]


def test_rename_and_delete(tmp_path):
    repo, _ = _repo_with_sales(tmp_path)
    chart = _save(repo)
    repo.rename_chart(chart.chart_id, "Regional revenue")
    assert [c.name for c in repo.charts()] == ["Regional revenue"]
    repo.delete_chart(chart.chart_id)
    assert repo.charts() == []


def test_unknown_chart_actions_fail_cleanly(tmp_path):
    repo, _ = _repo_with_sales(tmp_path)
    with pytest.raises(UnknownChartError):
        repo.open_chart("nope")
    with pytest.raises(UnknownChartError):
        repo.rename_chart("nope", "x")
    with pytest.raises(UnknownChartError):
        repo.delete_chart("nope")


# --------------------------------------------------------------------------- #
# Open = re-run the stored, validated SQL against CURRENT data
# --------------------------------------------------------------------------- #
def test_open_chart_reruns_the_query(tmp_path):
    repo, name = _repo_with_sales(tmp_path)
    chart = _save(repo)
    answer = repo.open_chart(chart.chart_id)
    rows = {r[0]: r[1] for r in answer.table.rows}
    assert rows == {"East": 60.0, "West": 90.0}
    assert answer.chart_type == "bar"
    assert answer.trust_trail is not None
    assert TOTALS_SQL in (answer.trust_trail.sql or "")


def test_open_chart_sees_refreshed_data(tmp_path):
    repo, name = _repo_with_sales(tmp_path)
    chart = _save(repo)
    repo.refresh(name, "sales.csv", b"region,amount\nEast,1\nWest,400\n")
    answer = repo.open_chart(chart.chart_id)
    rows = {r[0]: r[1] for r in answer.table.rows}
    assert rows["West"] == 400.0


def test_open_chart_whose_dataset_is_gone_fails_clearly(tmp_path):
    repo, name = _repo_with_sales(tmp_path)
    chart = _save(repo)
    repo.delete(name)
    with pytest.raises(ChartDataGoneError):
        repo.open_chart(chart.chart_id)
    repo.delete_chart(chart.chart_id)  # still deletable
    assert repo.charts() == []


# --------------------------------------------------------------------------- #
# Line inference: temporal first column -> line; categorical stays bar
# --------------------------------------------------------------------------- #
from analyst.api.qa import shape_answer  # noqa: E402
from analyst.domain.query import PlanAction, QueryPlan, ResultTable  # noqa: E402


def _plan(sql: str = "SELECT 1") -> QueryPlan:
    return QueryPlan(action=PlanAction.ANSWER, sql=sql, title="T")


def test_temporal_axis_infers_a_line_chart():
    result = ResultTable(
        columns=("month", "total"),
        rows=(("2026-01", 10.0), ("2026-02", 20.0), ("2026-03", 15.0)),
    )
    assert shape_answer(_plan(), result).chart_type == "line"


def test_full_dates_also_infer_line():
    result = ResultTable(
        columns=("day", "total"),
        rows=(("2026-01-01", 1.0), ("2026-01-02", 2.0)),
    )
    assert shape_answer(_plan(), result).chart_type == "line"


def test_categorical_axis_stays_bar():
    result = ResultTable(
        columns=("region", "total"), rows=(("East", 60.0), ("West", 90.0))
    )
    assert shape_answer(_plan(), result).chart_type == "bar"


# --------------------------------------------------------------------------- #
# Exports: full fidelity, all formats, via the engine
# --------------------------------------------------------------------------- #
import csv  # noqa: E402
import io  # noqa: E402

from analyst.engine.exports import export_dataset, export_query  # noqa: E402


def test_dataset_exports_all_formats_with_all_rows(tmp_path):
    repo, name = _repo_with_sales(tmp_path)
    out = {}
    for fmt in ("csv", "parquet", "xlsx"):
        path = tmp_path / f"out.{fmt}"
        export_dataset(repo.store, name, fmt, path)
        out[fmt] = path

    rows = list(csv.reader(io.StringIO(out["csv"].read_text())))
    assert rows[0] == ["region", "amount"] and len(rows) == 6

    import duckdb

    con = duckdb.connect()
    assert (
        con.execute(
            f"SELECT COUNT(*) FROM read_parquet('{out['parquet']}')"
        ).fetchone()[0]
        == 5
    )

    from openpyxl import load_workbook

    sheet = load_workbook(out["xlsx"]).active
    assert sheet.max_row == 6  # header + 5 rows


def test_query_export_is_never_display_capped(tmp_path):
    repo = _repo(tmp_path)
    body = "\n".join(f"r{i},{i}" for i in range(500))
    (rec,) = repo.ingest("big.csv", f"key,val\n{body}\n".encode())
    path = tmp_path / "q.csv"
    export_query(repo.store, f'SELECT * FROM "{rec.name}"', "csv", path)
    rows = list(csv.reader(io.StringIO(path.read_text())))
    assert len(rows) == 501  # header + ALL 500 rows, not the 200-row display cap


def test_export_unknown_format_rejected(tmp_path):
    repo, name = _repo_with_sales(tmp_path)
    with pytest.raises(ValueError):
        export_dataset(repo.store, name, "pdf", tmp_path / "x.pdf")


# --------------------------------------------------------------------------- #
# API routes: chart CRUD + open + exports
# --------------------------------------------------------------------------- #
from fastapi.testclient import TestClient  # noqa: E402

from analyst.api.app import create_app  # noqa: E402
from analyst.api.repository import FixtureRepository  # noqa: E402

SAVE_BODY = {
    "name": "Revenue by region",
    "question": "What is the total amount by region?",
    "sql": TOTALS_SQL,
    "chartType": "bar",
    "title": "Total amount by region",
    "datasets": ["sales.csv"],
}


def _client(tmp_path) -> TestClient:
    repo, _ = _repo_with_sales(tmp_path)
    return TestClient(create_app(repo))


def test_chart_crud_over_http(tmp_path):
    client = _client(tmp_path)
    saved = client.post("/api/charts", json=SAVE_BODY).json()
    assert saved["chartId"] == "revenue-by-region"
    assert [c["name"] for c in client.get("/api/charts").json()["charts"]] == [
        "Revenue by region"
    ]
    opened = client.get("/api/charts/revenue-by-region").json()
    assert opened["chartType"] == "bar"
    assert opened["trustTrail"]["sql"] == TOTALS_SQL
    assert ["East", 60.0] in opened["table"]["rows"] or ["East", 60] in opened["table"][
        "rows"
    ]
    client.patch("/api/charts/revenue-by-region", json={"name": "Regional revenue"})
    assert [c["name"] for c in client.get("/api/charts").json()["charts"]] == [
        "Regional revenue"
    ]
    assert client.delete("/api/charts/revenue-by-region").status_code == 204
    assert client.get("/api/charts").json()["charts"] == []


def test_unknown_chart_is_404_data_gone_is_409(tmp_path):
    client = _client(tmp_path)
    assert client.get("/api/charts/nope").status_code == 404
    client.post("/api/charts", json=SAVE_BODY)
    client.delete("/api/datasets/sales.csv")
    response = client.get("/api/charts/revenue-by-region")
    assert response.status_code == 409
    assert "gone" in response.json()["detail"].lower()


def test_dataset_export_route(tmp_path):
    client = _client(tmp_path)
    response = client.get("/api/datasets/sales.csv/export?format=csv")
    assert response.status_code == 200
    assert "attachment" in response.headers["content-disposition"]
    lines = response.text.strip().splitlines()
    assert lines[0] == "region,amount" and len(lines) == 6
    assert client.get("/api/datasets/nope/export?format=csv").status_code == 404
    assert client.get("/api/datasets/sales.csv/export?format=pdf").status_code == 400


def test_chart_export_route(tmp_path):
    client = _client(tmp_path)
    client.post("/api/charts", json=SAVE_BODY)
    response = client.get("/api/charts/revenue-by-region/export?format=csv")
    assert response.status_code == 200
    assert "West,90" in response.text.replace(".0", "")


def test_fixture_repo_supports_the_chart_flow():
    client = TestClient(create_app(FixtureRepository()))
    saved = client.post(
        "/api/charts", json={**SAVE_BODY, "name": "Regional revenue"}
    ).json()
    assert saved["chartId"] == "regional-revenue"
    opened = client.get("/api/charts/regional-revenue").json()
    assert opened["chartType"] in {"bar", "line"}
    assert opened["trustTrail"] is not None
