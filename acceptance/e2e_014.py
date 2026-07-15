"""Step handlers for feature 014 — charts & data exports.

Chart-lifecycle and export scenarios bind over the in-process seam (real
StoreRepository + chart service in the scenario tmp_path; restarts rebuild
the stack). The two workbench flows bind to Playwright against the fixtures
app. Deterministic — reopening a chart executes its stored query; no model
calls anywhere.

Bindings land slice by slice during implementation (CP5); unbound steps fail
the board explicitly with NOT YET IMPLEMENTED — the intended red.
"""

from __future__ import annotations

import csv
import io
import os
from typing import Any

import httpx

from acceptance.e2e_base import (
    _STACK,
    ScenarioContext,
    _e2e_fresh,
    _e2e_stack,
    expect_,
    make_registry,
)
from analyst.api.qa import shape_answer
from analyst.api.repository import StoreRepository
from analyst.domain.charts import ChartDataGoneError, UnknownChartError
from analyst.domain.query import PlanAction, QueryPlan
from analyst.engine.exports import export_dataset, export_query
from analyst.engine.query import run_select

step, run_step = make_registry()
_expect = expect_

__all__ = ["ScenarioContext", "run_step", "_e2e_stack", "_e2e_fresh"]


TOTALS_SQL = (
    'SELECT region, SUM(amount) AS total FROM "sales.csv" '
    "GROUP BY region ORDER BY total DESC"
)


def _state(ctx: ScenarioContext) -> dict[str, Any]:
    if not isinstance(ctx.data, dict):
        ctx.data = {}
    return ctx.data


def _repo(ctx: ScenarioContext) -> StoreRepository:
    state = _state(ctx)
    if "repo" not in state:
        state["repo"] = StoreRepository(str(ctx.tmp_path / "data"))
    return state["repo"]


def _restart(ctx: ScenarioContext) -> StoreRepository:
    state = _state(ctx)
    state["repo"] = StoreRepository(str(ctx.tmp_path / "data"))
    return state["repo"]


def _answer(ctx: ScenarioContext, sql: str, title: str):  # noqa: ANN202
    repo = _repo(ctx)
    plan = QueryPlan(action=PlanAction.ANSWER, sql=sql, title=title)
    return shape_answer(plan, run_select(repo.store, sql))


# --------------------------------------------------------------------------- #
# Givens
# --------------------------------------------------------------------------- #
@step(
    r'an ingested sales file where case variants of "East" carry amounts '
    r'10, 20 and 30 and "West" carries 40 and 50'
)
def given_sales_file(ctx: ScenarioContext) -> None:
    (rec,) = _repo(ctx).ingest(
        "sales.csv", b"region,amount\nEast,10\neast,20\nEAST,30\nWest,40\nWest,50\n"
    )
    _state(ctx)["dataset"] = rec.name


@step(r"an ingested file of monthly totals across six months")
def given_monthly_file(ctx: ScenarioContext) -> None:
    rows = "\n".join(f"2026-0{m},{m * 10}" for m in range(1, 7))
    (rec,) = _repo(ctx).ingest("monthly.csv", f"month,total\n{rows}\n".encode())
    _state(ctx)["dataset"] = rec.name


@step(r"an ingested file with more rows than the display cap")
def given_big_file(ctx: ScenarioContext) -> None:
    body = "\n".join(f"r{i},{i}" for i in range(500))
    (rec,) = _repo(ctx).ingest("big.csv", f"key,val\n{body}\n".encode())
    _state(ctx)["dataset"] = rec.name


@step(r"an answered question computing the total amount by region")
def given_answered_question(ctx: ScenarioContext) -> None:
    _state(ctx)["answer"] = _answer(ctx, TOTALS_SQL, "Total amount by region")


@step(r'a saved chart named "(?P<name>[^"]+)" computing the total amount by region')
def given_saved_chart(ctx: ScenarioContext, name: str) -> None:
    chart = _repo(ctx).save_chart(
        name=name,
        question="What is the total amount by region?",
        sql=TOTALS_SQL,
        chart_type="bar",
        title="Total amount by region",
        datasets=["sales.csv"],
    )
    _state(ctx)["chart_id"] = chart.chart_id


@step(r"the app runs offline with no AI features available")
def given_offline(ctx: ScenarioContext) -> None:
    os.environ.pop("ANALYST_CATALOG", None)
    os.environ.pop("ANALYST_CATALOG_CASSETTE", None)


@step(r'the proposed rule for column "(?P<column>[^"]+)" is approved')
def given_rule_approved(ctx: ScenarioContext, column: str) -> None:
    state = _state(ctx)
    state["repo"].approve_normalization(state["dataset"], f"norm:{column}")


# --------------------------------------------------------------------------- #
# Whens
# --------------------------------------------------------------------------- #
@step(r"the user's question computes the total by month")
def when_totals_by_month(ctx: ScenarioContext) -> None:
    sql = (
        'SELECT month, SUM(total) AS t FROM "monthly.csv" GROUP BY month ORDER BY month'
    )
    _state(ctx)["answer"] = _answer(ctx, sql, "Total by month")


@step(r"the user's question computes the total amount by region")
def when_totals_by_region(ctx: ScenarioContext) -> None:
    _state(ctx)["answer"] = _answer(ctx, TOTALS_SQL, "Total amount by region")


@step(r'the user saves the answer as a chart named "(?P<name>[^"]+)"')
def when_save_answer(ctx: ScenarioContext, name: str) -> None:
    state = _state(ctx)
    if "answer" in state:  # in-process seam
        answer = state["answer"]
        chart = _repo(ctx).save_chart(
            name=name,
            question="What is the total amount by region?",
            sql=answer.trust_trail.sql,
            chart_type=answer.chart_type,
            title=answer.chart_title or name,
            datasets=["sales.csv"],
        )
        state["chart_id"] = chart.chart_id
        return
    # browser flow (fixtures app)
    ctx.page.get_by_role("button", name="Save as chart").click()
    field = ctx.page.get_by_label("Chart name")
    field.fill(name)
    ctx.page.get_by_role("button", name="Confirm save chart").click()
    _expect()(ctx.page.get_by_text("Saved to Charts", exact=False)).to_be_visible()


@step(r"the user opens the saved chart")
def when_open_chart(ctx: ScenarioContext) -> None:
    state = _state(ctx)
    try:
        state["opened"] = state["repo"].open_chart(state["chart_id"])
        state["error"] = None
    except (ChartDataGoneError, UnknownChartError) as exc:
        state["opened"], state["error"] = None, exc


@step(r'the dataset is refreshed so "West" carries 400 instead')
def when_refresh_west(ctx: ScenarioContext) -> None:
    state = _state(ctx)
    state["repo"].refresh(
        state["dataset"], "sales.csv", b"region,amount\nEast,10\nWest,400\n"
    )


@step(r'the user renames the saved chart to "(?P<name>[^"]+)"')
def when_rename_chart(ctx: ScenarioContext, name: str) -> None:
    state = _state(ctx)
    state["repo"].rename_chart(state["chart_id"], name)


@step(r"the user deletes the saved chart")
def when_delete_chart(ctx: ScenarioContext) -> None:
    state = _state(ctx)
    state["repo"].delete_chart(state["chart_id"])


@step(r"the app restarts")
def when_app_restarts(ctx: ScenarioContext) -> None:
    _restart(ctx)


@step(r"the dataset is deleted")
def when_dataset_deleted(ctx: ScenarioContext) -> None:
    state = _state(ctx)
    state["repo"].delete(state["dataset"])


@step(r"the user opens a saved chart that does not exist")
def when_open_missing_chart(ctx: ScenarioContext) -> None:
    state = _state(ctx)
    try:
        state["repo"].open_chart("no-such-chart")
        state["error"] = None
    except UnknownChartError as exc:
        state["error"] = exc


@step(r"the user exports the chart's result as CSV and as Excel")
def when_export_chart_result(ctx: ScenarioContext) -> None:
    state = _state(ctx)
    charts = {c.chart_id: c for c in state["repo"].charts()}
    sql = charts[state["chart_id"]].sql
    csv_path = ctx.tmp_path / "chart.csv"
    xlsx_path = ctx.tmp_path / "chart.xlsx"
    export_query(state["repo"].store, sql, "csv", csv_path)
    export_query(state["repo"].store, sql, "xlsx", xlsx_path)
    state["exports"] = {"csv": csv_path, "xlsx": xlsx_path}


@step(r"the user exports the dataset in each of the three formats")
def when_export_dataset_all(ctx: ScenarioContext) -> None:
    state = _state(ctx)
    out = {}
    for fmt in ("csv", "parquet", "xlsx"):
        path = ctx.tmp_path / f"ds.{fmt}"
        export_dataset(state["repo"].store, state["dataset"], fmt, path)
        out[fmt] = path
    state["exports"] = out


@step(r"the user exports the dataset as CSV")
def when_export_dataset_csv(ctx: ScenarioContext) -> None:
    state = _state(ctx)
    path = ctx.tmp_path / "ds.csv"
    export_dataset(state["repo"].store, state["dataset"], "csv", path)
    state["exports"] = {"csv": path}


@step(r"the user exports a dataset that does not exist")
def when_export_missing_dataset(ctx: ScenarioContext) -> None:
    state = _state(ctx)
    try:
        export_dataset(
            state["repo"].store, "no-such-dataset", "csv", ctx.tmp_path / "x.csv"
        )
        state["error"] = None
    except KeyError as exc:
        state["error"] = exc


# --------------------------------------------------------------------------- #
# Thens
# --------------------------------------------------------------------------- #
@step(r"the answer presents as a line chart")
def then_line_chart(ctx: ScenarioContext) -> None:
    assert _state(ctx)["answer"].chart_type == "line"


@step(r"the answer presents as a bar chart")
def then_bar_chart(ctx: ScenarioContext) -> None:
    assert _state(ctx)["answer"].chart_type == "bar"


@step(r'the workspace lists a saved chart named "(?P<name>[^"]+)"')
def then_chart_listed(ctx: ScenarioContext, name: str) -> None:
    names = [c.name for c in _state(ctx)["repo"].charts()]
    assert name in names, names


@step(r"the workspace lists no saved charts")
def then_no_charts(ctx: ScenarioContext) -> None:
    assert _state(ctx)["repo"].charts() == []


@step(r'the chart shows "(?P<label>[^"]+)" totalling (?P<total>\d+)')
def then_chart_total(ctx: ScenarioContext, label: str, total: str) -> None:
    opened = _state(ctx)["opened"]
    rows = {str(r[0]): float(r[1]) for r in opened.table.rows}
    assert rows.get(label) == float(total), rows


@step(r"the chart carries a trust trail disclosing its query")
def then_chart_trust_trail(ctx: ScenarioContext) -> None:
    opened = _state(ctx)["opened"]
    assert opened.trust_trail is not None and opened.trust_trail.sql


@step(r'opening it shows "(?P<label>[^"]+)" totalling (?P<total>\d+)')
def then_open_and_check(ctx: ScenarioContext, label: str, total: str) -> None:
    state = _state(ctx)
    opened = state["repo"].open_chart(state["chart_id"])
    rows = {str(r[0]): float(r[1]) for r in opened.table.rows}
    assert rows.get(label) == float(total), rows


@step(r"the chart reports that its data is gone")
def then_data_gone(ctx: ScenarioContext) -> None:
    error = _state(ctx)["error"]
    assert isinstance(error, ChartDataGoneError), error
    assert "gone" in str(error).lower()


@step(r"the user can still delete the saved chart")
def then_still_deletable(ctx: ScenarioContext) -> None:
    state = _state(ctx)
    state["repo"].delete_chart(state["chart_id"])
    assert all(c.chart_id != state["chart_id"] for c in state["repo"].charts())


@step(r"each export's header names the result's columns")
def then_export_headers(ctx: ScenarioContext) -> None:
    exports = _state(ctx)["exports"]
    rows = list(csv.reader(io.StringIO(exports["csv"].read_text())))
    assert rows[0] == ["region", "total"], rows[0]
    from openpyxl import load_workbook

    sheet = load_workbook(exports["xlsx"]).active
    assert [c.value for c in next(sheet.iter_rows())] == ["region", "total"]


@step(r"each export's rows carry \"(?P<label>[^\"]+)\" with (?P<total>\d+)")
def then_export_rows(ctx: ScenarioContext, label: str, total: str) -> None:
    exports = _state(ctx)["exports"]
    rows = list(csv.reader(io.StringIO(exports["csv"].read_text())))
    assert [label, total] in [[r[0], r[1].removesuffix(".0")] for r in rows[1:]]
    from openpyxl import load_workbook

    sheet = load_workbook(exports["xlsx"]).active
    data = {str(r[0].value): float(r[1].value) for r in list(sheet.iter_rows())[1:]}
    assert data.get(label) == float(total)


@step(r"every export carries all (?P<n>\d+) rows of the dataset")
def then_exports_full(ctx: ScenarioContext, n: str) -> None:
    import duckdb

    exports = _state(ctx)["exports"]
    expected = int(n)
    rows = list(csv.reader(io.StringIO(exports["csv"].read_text())))
    assert len(rows) == expected + 1
    con = duckdb.connect()
    count = con.execute(
        f"SELECT COUNT(*) FROM read_parquet('{exports['parquet']}')"
    ).fetchone()[0]
    assert count == expected
    from openpyxl import load_workbook

    assert load_workbook(exports["xlsx"]).active.max_row == expected + 1


@step(r"the export's \"region\" values are only \"East\" and \"West\"")
def then_export_normalized(ctx: ScenarioContext) -> None:
    exports = _state(ctx)["exports"]
    rows = list(csv.reader(io.StringIO(exports["csv"].read_text())))
    assert {r[0] for r in rows[1:]} == {"East", "West"}


@step(r"the export carries every row while the on-screen table is capped")
def then_export_uncapped(ctx: ScenarioContext) -> None:
    state = _state(ctx)
    rows = list(csv.reader(io.StringIO(state["exports"]["csv"].read_text())))
    assert len(rows) == 501, len(rows)
    display = run_select(state["repo"].store, f'SELECT * FROM "{state["dataset"]}"')
    assert display.truncated and len(display.rows) < 500


@step(r"the action is rejected as not found")
def then_not_found(ctx: ScenarioContext) -> None:
    error = _state(ctx)["error"]
    assert isinstance(error, (UnknownChartError, KeyError)), error


# --------------------------------------------------------------------------- #
# Workbench flows (browser) — fixtures app
# --------------------------------------------------------------------------- #
@step(r"the analyst app is open in a browser")
def given_app_open(ctx: ScenarioContext) -> None:
    expect = _expect()
    ctx.page.goto(_STACK["web"])
    expect(ctx.page.get_by_text("Catalog", exact=True).first).to_be_visible()


@step(r"the user has an answered revenue question in the thread")
def given_answered_in_thread(ctx: ScenarioContext) -> None:
    expect = _expect()
    ctx.page.get_by_role("button", name="Query", exact=True).click()
    box = ctx.page.get_by_placeholder("Ask across all tables")
    box.fill("What is the revenue by region?")
    box.press("Enter")
    ctx.page.get_by_text("Two region columns are available.", exact=False).wait_for()
    ctx.page.get_by_role("button").filter(has_text="customer region").click()
    expect(
        ctx.page.get_by_text("East region generated the most", exact=False)
    ).to_be_visible()


@step(r"the user switches the answer's presentation to a line chart")
def when_switch_to_line(ctx: ScenarioContext) -> None:
    expect = _expect()
    ctx.page.get_by_role("button", name="Line", exact=True).click()
    expect(ctx.page.get_by_label("Line chart:", exact=False)).to_be_visible()


@step(r'the Charts area lists "(?P<name>[^"]+)"')
def then_charts_area_lists(ctx: ScenarioContext, name: str) -> None:
    expect = _expect()
    ctx.page.get_by_role("button", name="Charts", exact=True).click()
    expect(ctx.page.get_by_role("button", name=f"Open chart {name}")).to_be_visible()


@step(r'a saved chart named "(?P<name>[^"]+)" exists in the workspace')
def given_saved_chart_in_workspace(ctx: ScenarioContext, name: str) -> None:
    response = httpx.post(
        f"{ctx.api}/api/charts",
        json={
            "name": name,
            "question": "What is the revenue by region?",
            "sql": "SELECT region, SUM(amount) AS total FROM sales GROUP BY region",
            "chartType": "bar",
            "title": "Revenue by region",
        },
        timeout=10,
    )
    assert response.status_code == 200, response.text


@step(r'the user opens "(?P<name>[^"]+)" from the Charts area')
def when_open_from_charts_area(ctx: ScenarioContext, name: str) -> None:
    ctx.page.get_by_role("button", name="Charts", exact=True).click()
    ctx.page.get_by_role("button", name=f"Open chart {name}").click()


@step(r"the chart renders with its trust trail available")
def then_chart_renders_with_trail(ctx: ScenarioContext) -> None:
    expect = _expect()
    expect(ctx.page.get_by_text("Trust trail", exact=True)).to_be_visible()


@step(r"the user can switch its presentation to a table")
def then_switch_to_table(ctx: ScenarioContext) -> None:
    expect = _expect()
    ctx.page.get_by_role("button", name="Table", exact=True).click()
    expect(ctx.page.get_by_role("columnheader", name="region")).to_be_visible()
