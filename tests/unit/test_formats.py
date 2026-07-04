"""Unit tests for multi-format ingestion (Slice C)."""

import json

import pytest
from openpyxl import Workbook

from analyst.engine.reader import MalformedFileError, UnsupportedFormatError
from analyst.engine.store import DatasetStore
from analyst.service.ingestion import IngestionService


def _service(tmp_path):
    return IngestionService(DatasetStore(base_dir=tmp_path / "store"))


def test_tsv_is_ingested_like_csv(tmp_path):
    tsv = tmp_path / "data.tsv"
    tsv.write_text("id\tname\n1\talice\n2\tbob\n", encoding="utf-8")
    result = _service(tmp_path).ingest(tsv)
    assert [c.name for c in result.profile.columns] == ["id", "name"]
    assert result.profile.row_count == 2


def test_json_array_of_records_becomes_dataset(tmp_path):
    records = [{"id": i, "name": f"user{i}"} for i in range(50)]
    path = tmp_path / "orders.json"
    path.write_text(json.dumps(records), encoding="utf-8")
    result = _service(tmp_path).ingest(path)
    assert result.profile.row_count == 50
    assert {"id", "name"} == {c.name for c in result.profile.columns}


def test_nested_json_is_preserved_and_recorded(tmp_path):
    records = [
        {"id": 1, "shipping": {"city": "NYC", "zip": "10001"}},
        {"id": 2, "shipping": {"city": "LA", "zip": "90001"}},
    ]
    path = tmp_path / "orders.json"
    path.write_text(json.dumps(records), encoding="utf-8")
    result = _service(tmp_path).ingest(path)
    shipping = next(c for c in result.profile.columns if c.name == "shipping")
    assert shipping.is_nested is True
    # content preserved (as JSON text), not dropped
    rows = result.datasets[0].profile
    assert rows.row_count == 2


def test_excel_each_sheet_becomes_its_own_dataset(tmp_path):
    wb = Workbook()
    orders = wb.active
    orders.title = "orders"
    orders.append(["id", "total"])
    orders.append([1, 100])
    returns = wb.create_sheet("returns")
    returns.append(["order_id"])
    returns.append([1])
    path = tmp_path / "book.xlsx"
    wb.save(path)

    result = _service(tmp_path).ingest(path)
    names = {d.name for d in result.datasets}
    assert names == {"orders", "returns"}


def test_unsupported_format_is_rejected(tmp_path):
    path = tmp_path / "report.pdf"
    path.write_bytes(b"%PDF-1.4 not really")
    with pytest.raises(UnsupportedFormatError) as exc:
        _service(tmp_path).ingest(path)
    assert "CSV" in str(exc.value)


def test_malformed_excel_fails_cleanly(tmp_path):
    path = tmp_path / "broken.xlsx"
    path.write_bytes(b"this is not a real xlsx file")
    with pytest.raises(MalformedFileError):
        _service(tmp_path).ingest(path)


def test_M9_oversize_excel_is_rejected(tmp_path, monkeypatch):
    """M9: a workbook past the cell cap is rejected (too large), not OOM'd."""
    import pytest
    from openpyxl import Workbook

    from analyst.engine.excel import ExcelReader
    from analyst.engine.reader import FileTooLargeError

    monkeypatch.setenv("ANALYST_MAX_EXCEL_CELLS", "100")
    wb = Workbook()
    ws = wb.active
    for r in range(60):
        ws.append([f"c{c}" for c in range(5)])  # 300 cells > cap 100
    path = tmp_path / "big.xlsx"
    wb.save(path)
    with pytest.raises(FileTooLargeError):
        ExcelReader().sheets(path, tmp_path)
