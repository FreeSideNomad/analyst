"""Excel reader — expands a workbook into one normalized CSV per non-empty sheet.

Each sheet becomes its own dataset (AC-6). Supports both modern .xlsx (openpyxl)
and legacy .xls (xlrd). By converting sheets to CSV we reuse the delimited-file
materialization path.
"""

from __future__ import annotations

import csv
import os
from pathlib import Path

from openpyxl import load_workbook

from analyst.engine.reader import FileTooLargeError, MalformedFileError


def _sheet_has_content(rows: list[list[object]]) -> bool:
    return any(
        any(cell is not None and str(cell).strip() for cell in row) for row in rows
    )


class _TooManyCells(Exception):
    """Guard trip (M9): a workbook expands to more cells than the cap."""


def _max_cells() -> int:
    return int(os.environ.get("ANALYST_MAX_EXCEL_CELLS", str(5_000_000)))


def _xlsx_sheets(path: str) -> list[tuple[str, list[list[object]]]]:
    # SECURITY M9: cap total cells while streaming — a small, highly-compressed
    # .xlsx (zip bomb) or a sparse sheet with a cell at row 1,048,576 can expand
    # to enormous memory. Count as we go and abort past the cap.
    cap = _max_cells()
    workbook = load_workbook(filename=path, read_only=True, data_only=True)
    out: list[tuple[str, list[list[object]]]] = []
    seen = 0
    try:
        for sheet in workbook.worksheets:
            rows: list[list[object]] = []
            for row in sheet.iter_rows(values_only=True):
                seen += len(row)
                if seen > cap:
                    raise _TooManyCells
                rows.append(list(row))
            out.append((sheet.title, rows))
    finally:
        workbook.close()
    return out


def _xls_sheets(path: str) -> list[tuple[str, list[list[object]]]]:
    import xlrd

    cap = _max_cells()
    book = xlrd.open_workbook(path)
    total = sum(s.nrows * s.ncols for s in book.sheets())
    if total > cap:
        raise _TooManyCells
    return [
        (
            sheet.name,
            [
                [sheet.cell_value(r, c) for c in range(sheet.ncols)]
                for r in range(sheet.nrows)
            ],
        )
        for sheet in book.sheets()
    ]


class ExcelReader:
    """Reads an .xlsx/.xls workbook into per-sheet CSV files."""

    def sheets(
        self, path: str | os.PathLike[str], out_dir: Path
    ) -> list[tuple[str, Path]]:
        """Return (sheet_name, csv_path) for each non-empty sheet.

        Raises MalformedFileError if the workbook cannot be parsed.
        """
        reader = _xls_sheets if str(path).lower().endswith(".xls") else _xlsx_sheets
        try:
            sheets = reader(str(path))
        except _TooManyCells as exc:
            raise FileTooLargeError(
                f"The Excel workbook expands to more than {_max_cells()} cells — "
                "too large for this version."
            ) from exc
        except Exception as exc:  # openpyxl / xlrd raise several types on bad files
            raise MalformedFileError(
                f"The Excel file could not be read: {exc}"
            ) from exc

        out: list[tuple[str, Path]] = []
        for name, rows in sheets:
            if not _sheet_has_content(rows):
                continue
            csv_path = out_dir / f"sheet_{name}.csv"
            with csv_path.open("w", newline="", encoding="utf-8") as handle:
                writer = csv.writer(handle)
                for row in rows:
                    writer.writerow(["" if c is None else c for c in row])
            out.append((name, csv_path))
        return out
